import re
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from mymodule import *

def create_excel(input_filepath, output_filepath):
    """Creates and formats an Excel file based on extracted data."""

    def extract_patterns(content):
        match = re.search(r'assert\s+(.*?)\s+\$display\((.*?)\)', content)

        lines = content.splitlines()

        assertions = []
        comments = []

        for line in lines:
            line = line.strip()  # Ukloni prazne prostore na početku i kraju linije

            match = re.search(r'assert property \(@\(posedge iCLK\)\s+(.*?)\s+\$display\((.*?)\)', line)

            if match:
                assertions.append(match.group(1))  
                comments.append(match.group(2))   

        return comments, assertions

    content = read_file(input_filepath)
    comments, assertions = extract_patterns(content)

    print(f"Creating Excel file: {output_filepath}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assertions"

    headers = [
        "OPIS SVOJSTVA", "ASSERTION", "BODOVI"
    ]
    ws.append(headers)

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    colors = ["BAE1FF", "BAFFC9", "FFDFBA"]

    for i in range(len(comments)):
        ws.append([comments[i], assertions[i]])

    for col_num, col_title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)

        cell = ws[f'{col_letter}1']
        cell.font = Font(bold=True, size=12)

        max_length = max(len(col_title), max(len(str(cell.value)) for cell in ws[col_letter]))
        ws.column_dimensions[col_letter].width = max_length + 6  # Adding extra space for padding
        
        fill = PatternFill(start_color=colors[col_num - 1], end_color=colors[col_num - 1], fill_type="solid")
        for cell in ws[col_letter]:
            cell.alignment = alignment
            cell.border = border
            cell.fill = fill

    wb.save(output_filepath)
    print(f"Excel file saved successfully.")

def extract_logs(input_file, output_file, start_marker, end_marker):
    with open(input_file, 'r', encoding="utf-8") as infile, open(output_file, 'w', encoding="utf-8") as outfile:
        capturing = False
        for line in infile:
            if start_marker in line:
                capturing = True
            if capturing:
                outfile.write(line)
            if end_marker in line:
                capturing = False

def process_log_file(input_filename, output_filename):
    with open(input_filename, 'r', encoding="utf-8") as infile:
        # Učitavanje linija iz fajla
        lines = infile.readlines()
    
    # Filtriranje linija koje počinju sa # ASSERTION
    filtered_lines = [line[2:].strip() for line in lines if line.startswith('# ASSERTION')]

    # Uklanjanje duplikata
    unique_lines = list(set(filtered_lines))

    # Sortiranje linija prema broju nakon ASSERTION
    sorted_lines = sorted(unique_lines, key=lambda x: int(x.split()[1][:-1]))

    # Upisivanje rezultata u novi fajl
    with open(output_filename, 'w', encoding="utf-8") as outfile:
        for line in sorted_lines:
            outfile.write(line + '\n')

def filter_assertions(assertions_file, processed_file, output_file):
    # Učitavanje linija iz processed_run.log
    with open(processed_file, 'r', encoding="utf-8") as file:
        processed_lines = set(line.strip() for line in file.readlines())

    # Učitavanje linija iz assertions.txt i filtriranje
    with open(assertions_file, 'r', encoding="utf-8") as file:
        assertions_lines = file.readlines()

    # Filtriranje
    filtered_assertions = []
    for line in assertions_lines:
        if any(processed_line in line for processed_line in processed_lines):
            filtered_assertions.append(line)

    # Upisivanje rezultata u novi fajl
    with open(output_file, 'w', encoding="utf-8") as file:
        for line in filtered_assertions:
            file.write(line)

def main():
    log_file_dut = 'run_dut.log'
    log_file_tb = 'run_tb.log'
    processed_log_file_dut = 'processed_run_dut.log'
    processed_log_file_tb = 'processed_run_tb.log'
    assertions_file = 'assertions.txt'
    processed_assertions_file_dut = 'processed_assertions_dut.txt'
    processed_assertions_file_tb = 'processed_assertions_tb.txt'

    # Extract logs between 'TESTIRANJE_SIGNALA_TB::RUNNING' and 'TESTIRANJE_SIGNALA_TB::PASSED'
    extract_logs('run.log', log_file_tb, 'TESTIRANJE_SIGNALA_TB::RUNNING', 'TESTIRANJE_SIGNALA_TB::PASSED')
    # Extract logs between 'TESTIRANJE_SIGNALA_DUT::RUNNING' and 'TESTIRANJE_SIGNALA_DUT::PASSED'
    extract_logs('run.log', log_file_dut, 'TESTIRANJE_SIGNALA_DUT::RUNNING', 'TESTIRANJE_SIGNALA_DUT::PASSED')

    process_log_file(log_file_dut, processed_log_file_dut)
    process_log_file(log_file_tb, processed_log_file_tb)
    
    filter_assertions(assertions_file, processed_log_file_dut, processed_assertions_file_dut)
    filter_assertions(assertions_file, processed_log_file_tb, processed_assertions_file_tb)

    create_excel(processed_assertions_file_dut, 'tabela_svojstava_dut.xlsx')
    create_excel(processed_assertions_file_tb, 'tabela_svojstava_tb.xlsx')
if __name__ == '__main__':
    main()
